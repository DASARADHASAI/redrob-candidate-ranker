#!/usr/bin/env python3
"""
Convert the rank.py submission CSV into the .xlsx format required by the
hack2skill submission portal. Not part of the scored ranking step — purely a
format conversion for upload. Requires openpyxl (see requirements.txt).

Usage:
    python csv_to_xlsx.py --csv ./submission.csv --out ./submission.xlsx
"""
import argparse
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    rows = list(csv.reader(open(args.csv, encoding="utf-8")))
    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    for c, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E78")

    for r, row in enumerate(data, 2):
        cid, rank, score, reasoning = row
        ws.cell(row=r, column=1, value=cid).font = Font(name="Arial")
        rc = ws.cell(row=r, column=2, value=int(rank))
        rc.font = Font(name="Arial")
        rc.alignment = Alignment(horizontal="center")
        sc = ws.cell(row=r, column=3, value=float(score))
        sc.font = Font(name="Arial")
        sc.number_format = "0.0000"
        ws.cell(row=r, column=4, value=reasoning).font = Font(name="Arial")
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    for i, w in enumerate([16, 8, 10, 100], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(args.out)
    print(f"Wrote {args.out} ({len(data)} rows)")


if __name__ == "__main__":
    main()
